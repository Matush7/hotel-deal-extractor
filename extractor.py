"""
Hotel Deal Extractor — mini projekt pre Relaxos
================================================
Automaticky extrahuje štruktúrované dáta o akciových pobytoch
z neštruktúrovaných e-mailov od hotelov pomocou pattern matching.

Autor: [Tvoje meno]
Technológie: Python, Regex, openpyxl, JSON export
"""

import re
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sample_emails import SAMPLE_EMAILS


# ── Pomocné funkcie ───────────────────────────────────────────────────────────

def extract_hotel_name(text):
    patterns = [
        r'(Grand\s+Hotel\s+\w+)',
        r'(AquaCity\s+\w+(?:\s+\w+)?)',
        r'(Hotel\s+\w+(?:\s+\w+)?)',
        r'(Penzión\s+\w+(?:\s+\w+)?)',
        r'(Resort\s+\w+(?:\s+\w+)?)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def extract_stars(text):
    m = re.search(r'(\d)\s*\*{1,5}|\*{4,5}', text)
    if m:
        s = m.group(1)
        return int(s) if s and s.isdigit() else len(m.group(0).strip())
    return None


def extract_location(text):
    locations = [
        'Vysoké Tatry', 'Tatranská Lomnica', 'Nízke Tatry', 'Chopok',
        'Poprad', 'Žilina', 'Bratislava', 'Košice',
        'Štrbské Pleso', 'Liptovský Mikuláš',
    ]
    for loc in locations:
        if loc.lower() in text.lower():
            return loc
    return None


def extract_prices(text):
    discount = None
    m = re.search(r'(\d+)\s*%\s*(?:zľava|off|discount|-)', text, re.IGNORECASE)
    if m:
        discount = float(m.group(1))

    price_pairs = re.findall(r'(\d+(?:[.,]\d+)?)\s*EUR?', text, re.IGNORECASE)
    prices = [float(p.replace(',', '.')) for p in price_pairs if 10 < float(p.replace(',', '.')) < 2000]

    original, deal = None, None
    if len(prices) >= 2:
        original = max(prices[0], prices[1])
        deal = min(prices[0], prices[1])
        if not discount and original > 0:
            discount = round((1 - deal / original) * 100)
    elif len(prices) == 1 and discount:
        deal = prices[0]
        original = round(deal / (1 - discount / 100), 2)

    return original, deal, discount


def extract_dates(text):
    matches = re.findall(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})', text)
    dates = []
    for d, m, y in matches:
        try:
            dates.append(f"{y}-{int(m):02d}-{int(d):02d}")
        except:
            pass
    if len(dates) >= 2:
        return dates[0], dates[1]
    elif len(dates) == 1:
        return dates[0], None
    return None, None


def extract_room_type(block):
    patterns = [
        r'(Dvojlôžková izba\s*\w*)',
        r'(Apartmán\s*\w+(?:\s*\(\d\+\d\))?)',
        r'(Standard Double Room)',
        r'(Deluxe Room(?:\s+\w+)*)',
        r'(Business\s+\w+\s+Room)',
        r'(Superior\s+\w+(?:\s+\w+)?)',
        r'(2-lôžková\s+\w+(?:\s+\w+)?)',
        r'(Family\s+\w+)',
    ]
    for p in patterns:
        m = re.search(p, block, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def extract_includes(block):
    keywords = ['raňajky', 'polpenzia', 'breakfast', 'wellness', 'spa',
                'WiFi', 'parkovanie', 'parking', 'waterpark', 'masáž', 'massage',
                'deti', 'children']
    found = [kw for kw in keywords if kw.lower() in block.lower()]
    return ', '.join(found) if found else None


def extract_deal_name(block):
    patterns = [
        r'(?:AKCIA|PONUKA|BALÍČEK|PACKAGE|DEAL)\s*\d*[:\-]?\s*(.{5,60})',
        r'>>\s*(.{5,60}?)\s*<<',
        r'^(.{5,60}?)\n',
    ]
    for p in patterns:
        m = re.search(p, block.strip(), re.IGNORECASE | re.MULTILINE)
        if m:
            name = m.group(1).strip(' -:>\n')
            if len(name) > 4:
                return name
    return "Akciový pobyt"


def split_into_blocks(text):
    blocks = re.split(r'\n{2,}|(?=AKCIA\s*\d|>>|Izba:|Room:)', text)
    return [b.strip() for b in blocks if re.search(r'\d+\s*EUR?', b, re.IGNORECASE) and len(b) > 30]


# ── Extrakcia ─────────────────────────────────────────────────────────────────

def extract_deals_from_email(email):
    print(f"\n📧 Spracovávam: {email['subject']} ({email['from']})")

    text = email["body"]
    hotel_name = extract_hotel_name(text)
    stars = extract_stars(text)
    location = extract_location(text)

    blocks = split_into_blocks(text) or [text]

    deals = []
    for block in blocks:
        original, deal_price, discount = extract_prices(block)
        if not deal_price:
            continue

        date_from, date_to = extract_dates(block)
        room_type = extract_room_type(block)
        includes = extract_includes(block)
        deal_name = extract_deal_name(block)

        filled = sum([bool(hotel_name), bool(location), bool(original),
                      bool(deal_price), bool(discount), bool(date_from), bool(room_type)])
        confidence = round(filled / 7, 2)

        deals.append({
            "hotel_name": hotel_name or "",
            "hotel_stars": stars or "",
            "location": location or "",
            "deal_name": deal_name or "",
            "room_type": room_type or "",
            "date_from": date_from or "",
            "date_to": date_to or "",
            "original_price_eur": original or "",
            "deal_price_eur": deal_price or "",
            "discount_percent": discount or "",
            "includes": includes or "",
            "notes": "",
            "confidence": confidence,
            "source_email": email["from"],
            "extracted_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })

    print(f"   ✅ Nájdených {len(deals)} ponúk")
    for d in deals:
        conf = d.get('confidence', 0)
        emoji = "🟢" if conf >= 0.85 else "🟡" if conf >= 0.6 else "🔴"
        print(f"   {emoji} {d.get('deal_name')} | {d.get('deal_price_eur')}€ "
              f"({d.get('discount_percent')}% zľava) | istota: {conf:.0%}")

    return deals


# ── Export do Excel ───────────────────────────────────────────────────────────

COLUMNS = [
    ("Hotel", "hotel_name", 22),
    ("Hviezdy", "hotel_stars", 9),
    ("Lokalita", "location", 20),
    ("Názov akcie", "deal_name", 28),
    ("Typ izby", "room_type", 26),
    ("Platnosť od", "date_from", 13),
    ("Platnosť do", "date_to", 13),
    ("Cena pred (€)", "original_price_eur", 15),
    ("Cena po (€)", "deal_price_eur", 13),
    ("Zľava (%)", "discount_percent", 11),
    ("Zahŕňa", "includes", 30),
    ("Poznámky", "notes", 20),
    ("Istota", "confidence", 9),
    ("Zdroj e-mail", "source_email", 28),
    ("Extrahované", "extracted_at", 18),
]

def save_to_xlsx(all_deals, output_path):
    if not all_deals:
        print("⚠️  Žiadne dáta na uloženie.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Akciové pobyty"

    # Farby
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    thin = Border(
        left=Side(style='thin', color="CCCCCC"),
        right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin', color="CCCCCC"),
        bottom=Side(style='thin', color="CCCCCC"),
    )

    # Nadpis
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = "🏨 Relaxos — Extrahované akciové pobyty"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Hlavička
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # Dáta
    for row_idx, deal in enumerate(all_deals, start=3):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        conf = deal.get("confidence", 0)
        conf_fill = green_fill if conf >= 0.85 else yellow_fill if conf >= 0.6 else red_fill

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            val = deal.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if key == "confidence":
                cell.fill = conf_fill
                cell.value = f"{val:.0%}" if isinstance(val, float) else val
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key in ("original_price_eur", "deal_price_eur"):
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif key == "discount_percent":
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "hotel_stars":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 18

    # Zmraz prvé 2 riadky
    ws.freeze_panes = "A3"

    # Súhrn sheet
    ws2 = wb.create_sheet("Súhrn")
    ws2["A1"] = "📊 Súhrn extrakcie"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79")
    summary_data = [
        ("Celkový počet ponúk", len(all_deals)),
        ("Vysoká istota (≥85%)", len([d for d in all_deals if d.get("confidence", 0) >= 0.85])),
        ("Stredná istota (60-85%)", len([d for d in all_deals if 0.6 <= d.get("confidence", 0) < 0.85])),
        ("Nízka istota (<60%)", len([d for d in all_deals if d.get("confidence", 0) < 0.6])),
        ("", ""),
        ("Priemerná zľava", f"{sum(d['discount_percent'] for d in all_deals if d.get('discount_percent')) / len(all_deals):.1f}%"),
        ("Najvyššia zľava", f"{max((d['discount_percent'] for d in all_deals if d.get('discount_percent')), default=0)}%"),
        ("", ""),
        ("Vygenerované", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r, (label, val) in enumerate(summary_data, start=2):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16

    wb.save(output_path)
    print(f"\n💾 Excel uložený: {output_path}")


def save_to_json(all_deals, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_deals, f, ensure_ascii=False, indent=2)
    print(f"💾 JSON uložený: {output_path}")


def print_summary(all_deals):
    print("\n" + "="*60)
    print("📊 SÚHRN EXTRAKCIE")
    print("="*60)
    print(f"Celkový počet ponúk:     {len(all_deals)}")
    high = [d for d in all_deals if d.get("confidence", 0) >= 0.85]
    low  = [d for d in all_deals if d.get("confidence", 0) < 0.6]
    print(f"🟢 Vysoká istota (≥85%): {len(high)}")
    print(f"🔴 Nízka istota (<60%):  {len(low)}")
    if all_deals:
        disc = [d["discount_percent"] for d in all_deals if d.get("discount_percent")]
        if disc:
            print(f"\nPriemerná zľava:         {sum(disc)/len(disc):.1f}%")
            print(f"Najvyššia zľava:         {max(disc)}%")
    print("="*60)


# ── Spustenie ─────────────────────────────────────────────────────────────────

def main():
    print("🏨 Hotel Deal Extractor — Relaxos demo projekt")
    print("   (offline režim — pattern matching engine)")
    print("=" * 60)

    all_deals = []
    errors = []

    for email in SAMPLE_EMAILS:
        try:
            deals = extract_deals_from_email(email)
            all_deals.extend(deals)
        except Exception as e:
            print(f"   ❌ Chyba: {e}")
            errors.append(email["id"])

    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    save_to_xlsx(all_deals, f"{output_dir}/deals_{timestamp}.xlsx")
    save_to_json(all_deals, f"{output_dir}/deals_{timestamp}.json")
    print_summary(all_deals)

    if errors:
        print(f"\n❌ Nepodarilo sa spracovať: {errors}")

    print("\n✨ Hotovo! Excel je pripravený na import do katalógu Relaxos.")


if __name__ == "__main__":
    main()
